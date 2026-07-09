"""
Renders tests/data/test_cases.py (+ captured actual results, if any) into
tests/test_cases.xlsx.

Usage:
    uv run python -m tests.generate_test_case_workbook          # expected-only skeleton
    uv run pytest tests/ -m e2e                                 # runs tests, auto-regenerates with actual results
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from tests.data.test_cases import PIPELINE_TEST_CASES, RETRIEVAL_TEST_CASES

WORKBOOK_PATH = Path(__file__).parent / "test_cases.xlsx"
RESULTS_PATH = Path(__file__).parent / "artifacts" / "actual_results.json"

_HEADER_FILL = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_PASS_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
_FAIL_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
_WRAP = Alignment(wrap_text=True, vertical="top")

PIPELINE_HEADERS = [
    "ID", "Group", "Message (input)", "Expected Gender", "Expected Occasion Keywords",
    "Budget Max (VND)", "Required Roles (expected)", "Forbidden Categories (expected)",
    "Allowed Demographics (expected)", "Expect Dress-Only", "Min Days", "Notes",
    "Actual Summary", "Actual Days", "Actual Items", "Actual Total Price (day 1)",
    "Assertions", "Result",
]

RETRIEVAL_HEADERS = [
    "ID", "Group", "Query", "Target", "Expected Category (any of)",
    "Actual Hits", "Assertions", "Result",
]


def _style_header(ws, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP
    ws.freeze_panes = "A2"


def _autosize(ws, headers: list[str], width: int = 28) -> None:
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _format_items(items: list[dict]) -> str:
    lines = []
    for item in items:
        cats = ", ".join(item.get("categories", []))
        lines.append(
            f"[{item.get('target')}] {item.get('name')} "
            f"({cats}) - {item.get('target_demographic')} - "
            f"{item.get('base_price')} VND"
        )
    return "\n".join(lines)


def _format_assertions(assertions: list[dict]) -> str:
    return "\n".join(
        f"{'PASS' if a.get('passed') else 'FAIL'}: {a.get('name')} - {a.get('detail', '')}"
        for a in assertions
    )


def build_workbook(actual_results: dict[str, dict] | None = None) -> Path:
    if actual_results is None:
        actual_results = {}
        if RESULTS_PATH.exists():
            try:
                actual_results = json.loads(RESULTS_PATH.read_text(encoding="utf-8"))
            except Exception:
                actual_results = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Pipeline Test Cases"
    _style_header(ws, PIPELINE_HEADERS)
    _autosize(ws, PIPELINE_HEADERS)

    row = 2
    pass_count = 0
    fail_count = 0
    not_run_count = 0
    for case in PIPELINE_TEST_CASES:
        actual = actual_results.get(case["id"])
        required_roles_text = "\n".join(
            f"{r['role']}: {' / '.join(r['any_of'])}" for r in case["required_roles"]
        )
        result_text = "NOT RUN"
        if actual is not None:
            result_text = "PASS" if actual.get("passed") else "FAIL"
            if actual.get("passed"):
                pass_count += 1
            else:
                fail_count += 1
        else:
            not_run_count += 1

        values = [
            case["id"], case["group"], case["message"], case.get("expected_gender"),
            ", ".join(case.get("expected_occasion_keywords", [])),
            case.get("budget_max"), required_roles_text,
            ", ".join(case.get("forbidden_categories", [])),
            ", ".join(case.get("allowed_demographics", [])),
            case.get("expect_dress_only"), case.get("min_days"), case.get("notes"),
            actual.get("summary") if actual else "",
            json.dumps(actual.get("days"), ensure_ascii=False, indent=1) if actual else "",
            _format_items([i for d in actual.get("days", []) for i in d.get("items", [])]) if actual else "",
            actual.get("days", [{}])[0].get("total_price") if actual and actual.get("days") else None,
            _format_assertions(actual.get("assertions", [])) if actual else "",
            result_text,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = _WRAP
            if col == len(PIPELINE_HEADERS):
                if result_text == "PASS":
                    cell.fill = _PASS_FILL
                elif result_text == "FAIL":
                    cell.fill = _FAIL_FILL
        row += 1

    ws2 = wb.create_sheet("Retrieval Test Cases")
    _style_header(ws2, RETRIEVAL_HEADERS)
    _autosize(ws2, RETRIEVAL_HEADERS)
    row = 2
    for case in RETRIEVAL_TEST_CASES:
        actual = actual_results.get(case["id"])
        result_text = "NOT RUN"
        if actual is not None:
            result_text = "PASS" if actual.get("passed") else "FAIL"
            if actual.get("passed"):
                pass_count += 1
            else:
                fail_count += 1
        else:
            not_run_count += 1

        hits_text = ""
        if actual:
            hits_text = "\n".join(
                f"{h.get('product_id')} - {h.get('name')} ({', '.join(h.get('categories', []))})"
                for h in actual.get("hits", [])
            )
        values = [
            case["id"], case["group"], case["query"], case["target"],
            ", ".join(case["expected_category_any_of"]),
            hits_text,
            _format_assertions(actual.get("assertions", [])) if actual else "",
            result_text,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws2.cell(row=row, column=col, value=value)
            cell.alignment = _WRAP
            if col == len(RETRIEVAL_HEADERS):
                if result_text == "PASS":
                    cell.fill = _PASS_FILL
                elif result_text == "FAIL":
                    cell.fill = _FAIL_FILL
        row += 1

    ws3 = wb.create_sheet("Summary")
    ws3.append(["Total test cases", len(PIPELINE_TEST_CASES) + len(RETRIEVAL_TEST_CASES)])
    ws3.append(["Passed", pass_count])
    ws3.append(["Failed", fail_count])
    ws3.append(["Not run yet", not_run_count])
    for r in range(1, 5):
        ws3.cell(row=r, column=1).font = Font(bold=True)
    ws3.column_dimensions["A"].width = 24

    wb.save(WORKBOOK_PATH)
    return WORKBOOK_PATH


if __name__ == "__main__":
    path = build_workbook()
    print(f"Wrote {path}")
